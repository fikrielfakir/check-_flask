import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import tempfile
import os
from pathlib import Path
import logging

class ExcelManager:
    def __init__(self):
        self.upload_dir = Path("data/excel")
        self.upload_dir.mkdir(parents=True, exist_ok=True)
        
        # Excel headers - Updated to include deposit bank field
        self.headers = [
            "Date d'émission",
            "Type de Règlement", 
            "Numéro du chèque",
            "Banque/Agence",
            "Banque de dépôts - Agence",
            "Client",
            "Nom du déposant",
            "Montant",
            "Devise",
            "Date d'échéance",
            "Date de Création",
            "Statut",
            "N° Facture",
            "Date de facture",
            "Notes"
        ]
        
        # Month names in French
        self.month_names = [
            "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
            "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
        ]
        
        logging.info("Excel Manager initialized with optimized sync")
    
    def get_excel_filename(self, year):
        """Get Excel filename for a specific year"""
        return self.upload_dir / f"cheques_{year}.xlsx"
    
    def ensure_file_exists(self, year):
        """Ensure Excel file exists for the year, create if not"""
        filename = self.get_excel_filename(year)
        
        if not filename.exists():
            logging.info(f"Creating new Excel file for year {year}")
            self.create_yearly_file(year)
        
        return filename
    
    def create_yearly_file(self, year):
        """Create a new yearly Excel file with monthly sheets"""
        filename = self.get_excel_filename(year)
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Create 12 monthly sheets
        for month_name in self.month_names:
            sheet = workbook.create_sheet(title=month_name)
            self._setup_sheet_headers(sheet)
        
        workbook.save(filename)
        logging.info(f"Created yearly Excel file: {filename}")
        return filename
    
    def _setup_sheet_headers(self, sheet):
        """Setup headers and formatting for a sheet"""
        # Add headers
        for col, header in enumerate(self.headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            
            # Header formatting
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        
        # Set column widths for better readability
        column_widths = [12, 10, 15, 25, 25, 20, 12, 8, 12, 12, 12, 15, 12, 30]
        for col, width in enumerate(column_widths, 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # Freeze header row
        sheet.freeze_panes = "A2"
    
    def add_or_update_cheque(self, cheque):
        """Add or update a cheque in Excel - OPTIMIZED VERSION"""
        try:
            # Determine year and month from due date
            year = cheque.due_date.year
            month = cheque.due_date.month
            
            # Ensure Excel file exists
            filename = self.ensure_file_exists(year)
            
            # Load workbook
            workbook = openpyxl.load_workbook(filename)
            
            # Get or create month sheet
            month_name = self.month_names[month - 1]
            if month_name not in workbook.sheetnames:
                sheet = workbook.create_sheet(title=month_name)
                self._setup_sheet_headers(sheet)
            else:
                sheet = workbook[month_name]
            
            # Check for existing cheque
            existing_row = self._find_existing_cheque(sheet, cheque.cheque_number, cheque.branch.bank.name)
            
            # Prepare cheque data for Excel
            cheque_data = self._prepare_cheque_data(cheque)
            
            if existing_row:
                # Update existing row
                self._write_cheque_row(sheet, existing_row, cheque_data)
                logging.info(f"Updated existing cheque {cheque.cheque_number} in Excel")
            else:
                # Add new row
                next_row = self._get_next_empty_row(sheet)
                self._write_cheque_row(sheet, next_row, cheque_data)
                logging.info(f"Added new cheque {cheque.cheque_number} to Excel")
            
            # Apply formatting and save
            self._apply_row_formatting(sheet)
            workbook.save(filename)
            workbook.close()
            
            return True
            
        except Exception as e:
            logging.error(f"Error syncing cheque to Excel: {str(e)}")
            return False
    
    def _prepare_cheque_data(self, cheque):
        """Prepare cheque data in the correct order for Excel columns"""
        return [
            cheque.issue_date.strftime('%d/%m/%Y') if cheque.issue_date else '',
            cheque.payment_type or 'CHQ',
            cheque.cheque_number or '',
            f"{cheque.branch.bank.name} - {cheque.branch.name}" if cheque.branch else '',
            f"{cheque.deposit_branch.bank.name} - {cheque.deposit_branch.name}" if cheque.deposit_branch else '',
            cheque.client.name if cheque.client else '',
            cheque.depositor_name or '',
            float(cheque.amount),
            cheque.currency or 'MAD',
            cheque.due_date.strftime('%d/%m/%Y') if cheque.due_date else '',
            cheque.created_date.strftime('%d/%m/%Y') if cheque.created_date else '',
            cheque.status or 'EN ATTENTE',
            cheque.invoice_number or '',
            cheque.invoice_date.strftime('%d/%m/%Y') if cheque.invoice_date else '',
            cheque.notes or '' ''
        ]
    
    def _write_cheque_row(self, sheet, row_num, data):
        """Write cheque data to a specific row"""
        for col, value in enumerate(data, 1):
            cell = sheet.cell(row=row_num, column=col, value=value)
            
            # Apply borders
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            
            # Format amount column (column 7)
            if col == 7 and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
            
            # Center align certain columns
            if col in [2, 8, 9, 10, 11]:  # Type, Currency, Dates, Status
                cell.alignment = Alignment(horizontal="center")
    
    def _apply_row_formatting(self, sheet):
        """Apply alternating row colors and formatting"""
        # Light blue fill for alternating rows
        light_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
        
        # Apply alternating row colors (skip header)
        for row in range(2, sheet.max_row + 1):
            if row % 2 == 0:
                for col in range(1, len(self.headers) + 1):
                    cell = sheet.cell(row=row, column=col)
                    if not cell.fill.start_color.rgb or cell.fill.start_color.rgb == '00000000':
                        cell.fill = light_fill
    
    def _find_existing_cheque(self, sheet, cheque_number, bank_name):
        """Find existing cheque row by number and bank"""
        if not cheque_number:
            return None
        
        for row in range(2, sheet.max_row + 1):
            row_number = sheet.cell(row=row, column=3).value  # Numéro du chèque
            row_bank = sheet.cell(row=row, column=4).value    # Banque/Agence
            
            if (str(row_number or '').strip() == str(cheque_number).strip() and 
                row_bank and bank_name in str(row_bank)):
                return row
        return None
    
    def _get_next_empty_row(self, sheet):
        """Get next empty row number"""
        return sheet.max_row + 1
    
    def remove_cheque_from_excel(self, cheque_number, bank_name, year=None):
        """Remove a cheque from Excel file"""
        try:
            if year:
                years_to_check = [year]
            else:
                # Check current and previous year
                current_year = datetime.now().year
                years_to_check = [current_year, current_year - 1]
            
            for check_year in years_to_check:
                filename = self.get_excel_filename(check_year)
                if not filename.exists():
                    continue
                
                workbook = openpyxl.load_workbook(filename)
                
                for month_name in self.month_names:
                    if month_name not in workbook.sheetnames:
                        continue
                    
                    sheet = workbook[month_name]
                    existing_row = self._find_existing_cheque(sheet, cheque_number, bank_name)
                    
                    if existing_row:
                        sheet.delete_rows(existing_row)
                        workbook.save(filename)
                        workbook.close()
                        logging.info(f"Removed cheque {cheque_number} from Excel")
                        return True
                
                workbook.close()
            
            return False
            
        except Exception as e:
            logging.error(f"Error removing cheque from Excel: {str(e)}")
            return False
    
    def export_cheques(self, cheques, date_from=None, date_to=None):
        """Export cheques to Excel file"""
        # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.close()
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Export Chèques"
        
        # Setup headers
        self._setup_sheet_headers(sheet)
        
        # Add data
        for idx, cheque in enumerate(cheques, 2):
            data = self._prepare_cheque_data(cheque)
            self._write_cheque_row(sheet, idx, data)
        
        # Apply formatting
        self._apply_row_formatting(sheet)
        
        workbook.save(temp_file.name)
        return temp_file.name
    
    def get_file_statistics(self, year):
        """Get statistics for a yearly Excel file"""
        filename = self.get_excel_filename(year)
        
        if not filename.exists():
            return None
        
        try:
            workbook = openpyxl.load_workbook(filename, read_only=True)
            stats = {
                'year': year,
                'total_cheques': 0,
                'monthly_breakdown': {},
                'file_size': filename.stat().st_size,
                'last_modified': datetime.fromtimestamp(filename.stat().st_mtime)
            }
            
            for month_name in workbook.sheetnames:
                sheet = workbook[month_name]
                cheque_count = max(0, sheet.max_row - 1)  # Subtract header row
                stats['monthly_breakdown'][month_name] = cheque_count
                stats['total_cheques'] += cheque_count
            
            workbook.close()
            return stats
            
        except Exception as e:
            logging.error(f"Error getting file statistics: {str(e)}")
            return None