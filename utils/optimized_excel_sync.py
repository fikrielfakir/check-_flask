"""
Optimized Excel synchronization system with sheet/row tracking.
Implements the requirements from the attached file for robust Excel sync.
"""

import logging
import os
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook, Workbook
from models import db, ChequeExcelMapping

class OptimizedExcelSync:
    """Optimized Excel synchronization with persistent tracking"""
    
    def __init__(self, excel_folder_path):
        self.excel_folder = Path(excel_folder_path)
        self.excel_folder.mkdir(parents=True, exist_ok=True)
        self.logger = logging.getLogger(__name__)
        
        # Excel headers with deposit bank field
        self.headers = [
            "Date d'émission",
            "Type de Règlement", 
            "Numéro du chèque",
            "Banque/Agence",
            "Banque de dépôts - Agence",
            "Client",
            "Nom du déposant",
            "Montant",
            "Devise",
            "Date d'échéance",
            "Date de Création",
            "Statut",
            "N° Facture",
            "Date de facture",
            "Notes"
        ]
        
        # Month names in French
        self.month_names = [
            "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
            "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
        ]
    
    def sync_cheque(self, cheque, operation='create'):
        """
        Main synchronization method with tracking
        
        Args:
            cheque: Cheque model instance
            operation: 'create', 'update', or 'delete'
            
        Returns:
            bool: Success status
        """
        try:
            if operation == 'delete':
                return self._handle_cheque_deletion(cheque)
            else:
                return self._handle_cheque_upsert(cheque, operation)
                
        except Exception as e:
            self.logger.error(f"Error in sync_cheque: {str(e)}")
            return False
    
    def _handle_cheque_upsert(self, cheque, operation):
        """Handle cheque creation or update with tracking"""
        try:
            # Check if we have existing mapping
            existing_mapping = ChequeExcelMapping.query.filter_by(cheque_id=cheque.id).first()
            
            if existing_mapping and operation == 'update':
                # Update existing cheque in the same sheet/row
                return self._update_existing_cheque(cheque, existing_mapping)
            else:
                # Create new cheque entry
                return self._create_new_cheque(cheque)
                
        except Exception as e:
            self.logger.error(f"Error in _handle_cheque_upsert: {str(e)}")
            return False
    
    def _create_new_cheque(self, cheque):
        """Create new cheque entry with tracking"""
        try:
            # Determine target file and sheet
            year = cheque.due_date.year
            month = cheque.due_date.month
            month_name = self.month_names[month - 1]
            
            # Get or create Excel file
            filepath = self.excel_folder / f"cheques_{year}.xlsx"
            workbook, worksheet = self._ensure_workbook_and_sheet(filepath, month_name)
            
            # Find next available row
            next_row = worksheet.max_row + 1
            
            # Write cheque data
            cheque_data = self._prepare_cheque_data(cheque)
            for col_idx, value in enumerate(cheque_data, 1):
                worksheet.cell(row=next_row, column=col_idx, value=value)
            
            # Apply formatting
            self._apply_formatting(worksheet, next_row)
            
            # Save workbook
            workbook.save(filepath)
            workbook.close()
            
            # Create mapping record
            mapping = ChequeExcelMapping(
                cheque_id=cheque.id,
                excel_file_path=str(filepath),
                sheet_name=month_name,
                row_number=next_row
            )
            db.session.add(mapping)
            db.session.commit()
            
            self.logger.info(f"Created new cheque {cheque.id} in {filepath}, sheet {month_name}, row {next_row}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error creating new cheque: {str(e)}")
            db.session.rollback()
            return False
    
    def _update_existing_cheque(self, cheque, mapping):
        """Update existing cheque in its original location"""
        try:
            filepath = Path(mapping.excel_file_path)
            if not filepath.exists():
                self.logger.warning(f"Excel file not found: {filepath}")
                # File doesn't exist, recreate it
                return self._create_new_cheque(cheque)
            
            # Load workbook and sheet
            workbook = load_workbook(filepath)
            if mapping.sheet_name not in workbook.sheetnames:
                self.logger.warning(f"Sheet {mapping.sheet_name} not found in {filepath}")
                workbook.close()
                return self._create_new_cheque(cheque)
            
            worksheet = workbook[mapping.sheet_name]
            
            # Update the specific row
            cheque_data = self._prepare_cheque_data(cheque)
            for col_idx, value in enumerate(cheque_data, 1):
                worksheet.cell(row=mapping.row_number, column=col_idx, value=value)
            
            # Apply formatting
            self._apply_formatting(worksheet, mapping.row_number)
            
            # Save workbook
            workbook.save(filepath)
            workbook.close()
            
            # Update mapping timestamp
            mapping.updated_at = datetime.utcnow()
            db.session.commit()
            
            self.logger.info(f"Updated cheque {cheque.id} in {filepath}, sheet {mapping.sheet_name}, row {mapping.row_number}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error updating existing cheque: {str(e)}")
            db.session.rollback()
            return False
    
    def _handle_cheque_deletion(self, cheque):
        """Delete cheque from Excel and remove mapping"""
        try:
            mapping = ChequeExcelMapping.query.filter_by(cheque_id=cheque.id).first()
            if not mapping:
                self.logger.info(f"No Excel mapping found for cheque {cheque.id}")
                return True
            
            filepath = Path(mapping.excel_file_path)
            if not filepath.exists():
                self.logger.warning(f"Excel file not found: {filepath}")
                # Remove the mapping since file doesn't exist
                db.session.delete(mapping)
                db.session.commit()
                return True
            
            # Load workbook and delete row
            workbook = load_workbook(filepath)
            if mapping.sheet_name in workbook.sheetnames:
                worksheet = workbook[mapping.sheet_name]
                worksheet.delete_rows(mapping.row_number)
                workbook.save(filepath)
                
                # Update row numbers for subsequent mappings
                self._update_row_numbers_after_deletion(mapping)
            
            workbook.close()
            
            # Remove mapping
            db.session.delete(mapping)
            db.session.commit()
            
            self.logger.info(f"Deleted cheque {cheque.id} from Excel")
            return True
            
        except Exception as e:
            self.logger.error(f"Error deleting cheque: {str(e)}")
            db.session.rollback()
            return False
    
    def _update_row_numbers_after_deletion(self, deleted_mapping):
        """Update row numbers in mappings after a row deletion"""
        try:
            # Update all mappings in the same file/sheet with row numbers > deleted row
            mappings_to_update = ChequeExcelMapping.query.filter(
                ChequeExcelMapping.excel_file_path == deleted_mapping.excel_file_path,
                ChequeExcelMapping.sheet_name == deleted_mapping.sheet_name,
                ChequeExcelMapping.row_number > deleted_mapping.row_number
            ).all()
            
            for mapping in mappings_to_update:
                mapping.row_number -= 1
                mapping.updated_at = datetime.utcnow()
            
            db.session.commit()
            self.logger.info(f"Updated {len(mappings_to_update)} row numbers after deletion")
            
        except Exception as e:
            self.logger.error(f"Error updating row numbers: {str(e)}")
            db.session.rollback()
    
    def _ensure_workbook_and_sheet(self, filepath, sheet_name):
        """Ensure workbook and sheet exist"""
        if filepath.exists():
            workbook = load_workbook(filepath)
        else:
            workbook = Workbook()
            # Remove default sheet
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
        
        if sheet_name not in workbook.sheetnames:
            worksheet = workbook.create_sheet(title=sheet_name)
            # Add headers
            for col_idx, header in enumerate(self.headers, 1):
                worksheet.cell(row=1, column=col_idx, value=header)
            self._apply_header_formatting(worksheet)
        else:
            worksheet = workbook[sheet_name]
        
        return workbook, worksheet
    
    def _prepare_cheque_data(self, cheque):
        """Prepare cheque data for Excel"""
        return [
            cheque.issue_date.strftime('%d/%m/%Y') if cheque.issue_date else '',
            cheque.payment_type or 'CHQ',
            cheque.cheque_number or '',
            f"{cheque.branch.bank.name} - {cheque.branch.name}" if cheque.branch else '',
            f"{cheque.deposit_branch.bank.name} - {cheque.deposit_branch.name}" if cheque.deposit_branch else '',
            cheque.client.name if cheque.client else '',
            cheque.depositor_name or '',
            float(cheque.amount),
            cheque.currency or 'MAD',
            cheque.due_date.strftime('%d/%m/%Y') if cheque.due_date else '',
            cheque.created_date.strftime('%d/%m/%Y') if cheque.created_date else '',
            cheque.status or 'EN ATTENTE',
            cheque.invoice_number or '',
            cheque.invoice_date.strftime('%d/%m/%Y') if cheque.invoice_date else '',
            cheque.notes or ''
        ]
    
    def _apply_header_formatting(self, worksheet):
        """Apply formatting to header row"""
        from openpyxl.styles import Font, Alignment, PatternFill
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col in range(1, len(self.headers) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    def _apply_formatting(self, worksheet, row_number):
        """Apply formatting to data row"""
        from openpyxl.styles import Alignment
        
        # Center align certain columns (dates, amounts, status)
        center_cols = [1, 2, 8, 9, 10, 11, 12, 13]  # Date and status columns
        for col in center_cols:
            cell = worksheet.cell(row=row_number, column=col)
            cell.alignment = Alignment(horizontal="center")
    
    def batch_sync_all_cheques(self, cheques_query=None):
        """Batch synchronize all cheques"""
        from models import Cheque
        
        if cheques_query is None:
            cheques = Cheque.query.all()
        else:
            cheques = cheques_query.all()
        
        results = {
            'total': len(cheques),
            'successful': 0,
            'failed': 0,
            'errors': []
        }
        
        for cheque in cheques:
            try:
                if self.sync_cheque(cheque, 'create'):
                    results['successful'] += 1
                else:
                    results['failed'] += 1
            except Exception as e:
                results['failed'] += 1
                results['errors'].append(f"Cheque {cheque.id}: {str(e)}")
        
        self.logger.info(f"Batch sync completed: {results['successful']}/{results['total']} successful")
        return results
    
    def verify_integrity(self, year=None):
        """Verify Excel-database integrity"""
        try:
            from models import Cheque
            
            if year is None:
                year = datetime.now().year
            
            # Count cheques in database for the year
            db_count = Cheque.query.filter(
                db.extract('year', Cheque.due_date) == year
            ).count()
            
            # Count mappings for the year
            mapping_count = ChequeExcelMapping.query.filter(
                ChequeExcelMapping.excel_file_path.like(f'%cheques_{year}.xlsx')
            ).count()
            
            return {
                'year': year,
                'database_count': db_count,
                'mapping_count': mapping_count,
                'in_sync': db_count == mapping_count,
                'difference': abs(db_count - mapping_count)
            }
            
        except Exception as e:
            self.logger.error(f"Error verifying integrity: {str(e)}")
            return {'error': str(e)}