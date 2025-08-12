"""
utils/enhanced_excel_sync.py
Enhanced Excel synchronization system with COMPREHENSIVE duplicate prevention.
"""

import logging
import os
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from models import db, ChequeExcelMapping

class EnhancedExcelSync:
    """Enhanced Excel synchronization with COMPREHENSIVE duplicate prevention"""
    
    def __init__(self, excel_folder_path):
        self.excel_folder = Path(excel_folder_path)
        self.excel_folder.mkdir(parents=True, exist_ok=True)
        self.logger = logging.getLogger(__name__)
        
        # Excel headers with deposit bank field
        self.headers = [
            "Date réc",
            "Type rég", 
            "N° doc",
            "Bq/Agce",
            "Bq dép. - Agce",
            "Client",
            "Nom dép",
            "Mont",
            "Devise",
            "Date éch",
            "Date créat",
            "Statut",
            "N° Facture",
            "Date fact",
            "Notes"
        ]
        
        # Month names in French
        self.month_names = [
            "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
            "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
        ]

    def sync_cheque(self, cheque, operation='create'):
        """
        Main synchronization method with ENHANCED duplicate checking
        
        Args:
            cheque: Cheque model instance
            operation: 'create', 'update', or 'delete'
            
        Returns:
            bool: Success status
        """
        try:
            if operation == 'delete':
                return self._handle_cheque_deletion(cheque)
            else:
                return self._handle_cheque_upsert_enhanced(cheque, operation)
                
        except Exception as e:
            self.logger.error(f"Error in sync_cheque: {str(e)}")
            return False

    def _handle_cheque_upsert_enhanced(self, cheque, operation):
        """Enhanced handle cheque creation or update with COMPREHENSIVE duplicate checking"""
        try:
            # Step 1: Check database mapping first
            existing_mapping = ChequeExcelMapping.query.filter_by(cheque_id=cheque.id).first()
            
            # Step 2: COMPREHENSIVE duplicate check in Excel sheets
            excel_duplicates = self._find_all_cheque_duplicates_in_excel(cheque)
            
            # Step 3: Handle based on findings
            if existing_mapping:
                # We have a database mapping
                if self._validate_mapping_location(existing_mapping):
                    # Valid mapping exists - update in place and remove any other duplicates
                    success = self._update_existing_cheque(cheque, existing_mapping)
                    if success and excel_duplicates:
                        self._remove_duplicate_entries(excel_duplicates, existing_mapping)
                    return success
                else:
                    # Invalid mapping - clean up and handle duplicates
                    self._cleanup_invalid_mapping(existing_mapping)
                    return self._handle_duplicates_and_create_or_update(cheque, excel_duplicates)
            
            elif excel_duplicates:
                # No database mapping but duplicates exist in Excel
                return self._handle_duplicates_and_create_or_update(cheque, excel_duplicates)
            
            else:
                # No mapping and no Excel duplicates - safe to create new
                return self._create_new_cheque(cheque)
                
        except Exception as e:
            self.logger.error(f"Error in _handle_cheque_upsert_enhanced: {str(e)}")
            return False

    def _find_all_cheque_duplicates_in_excel(self, cheque):
        """
        COMPREHENSIVE search for ALL duplicate entries of this cheque in Excel files
        Returns: list of dicts with file_path, sheet_name, row_number for each duplicate found
        """
        duplicates = []
        
        try:
            if not cheque.cheque_number or not cheque.cheque_number.strip():
                return duplicates
            
            cheque_number = cheque.cheque_number.strip()
            year = cheque.due_date.year
            filepath = self.excel_folder / f"cheques_{year}.xlsx"
            
            if not filepath.exists():
                return duplicates
            
            workbook = load_workbook(filepath, read_only=True)
            
            # Check ALL sheets for duplicates
            for sheet_name in workbook.sheetnames:
                try:
                    worksheet = workbook[sheet_name]
                    
                    # Search from row 2 (skip headers) through all rows
                    for row_num in range(2, worksheet.max_row + 1):
                        try:
                            excel_cheque_num = worksheet.cell(row=row_num, column=3).value  # Column 3 is "N° doc"
                            
                            if excel_cheque_num and str(excel_cheque_num).strip() == cheque_number:
                                # Found a match - get additional details for verification
                                excel_client = worksheet.cell(row=row_num, column=6).value or ""  # Client column
                                excel_amount = worksheet.cell(row=row_num, column=8).value  # Amount column
                                
                                duplicate_info = {
                                    'file_path': filepath,
                                    'sheet_name': sheet_name,
                                    'row_number': row_num,
                                    'cheque_number': str(excel_cheque_num).strip(),
                                    'client_name': str(excel_client).strip(),
                                    'amount': excel_amount,
                                    'is_exact_match': self._is_exact_cheque_match(cheque, worksheet, row_num)
                                }
                                duplicates.append(duplicate_info)
                                
                                self.logger.info(f"Found duplicate cheque {cheque_number} in {sheet_name} at row {row_num}")
                                
                        except Exception as e:
                            self.logger.warning(f"Error checking row {row_num} in sheet {sheet_name}: {str(e)}")
                            continue
                            
                except Exception as e:
                    self.logger.error(f"Error checking sheet {sheet_name}: {str(e)}")
                    continue
            
            workbook.close()
            
            if duplicates:
                self.logger.warning(f"Found {len(duplicates)} duplicate entries for cheque {cheque_number}")
            
            return duplicates
            
        except Exception as e:
            self.logger.error(f"Error searching for duplicates: {str(e)}")
            return duplicates

    def _is_exact_cheque_match(self, cheque, worksheet, row_num):
        """
        Check if the Excel row is an exact match for the given cheque
        """
        try:
            # Get values from Excel row
            excel_cheque_num = str(worksheet.cell(row=row_num, column=3).value or "").strip()
            excel_client = str(worksheet.cell(row=row_num, column=6).value or "").strip()
            excel_amount = worksheet.cell(row=row_num, column=8).value
            
            # Get cheque values
            cheque_number = cheque.cheque_number.strip() if cheque.cheque_number else ""
            client_name = cheque.client.name.strip() if cheque.client else ""
            cheque_amount = float(cheque.amount) if cheque.amount else 0.0
            
            # Convert excel amount to float for comparison
            try:
                excel_amount_float = float(excel_amount) if excel_amount else 0.0
            except (ValueError, TypeError):
                excel_amount_float = 0.0
            
            # Check for exact match
            return (excel_cheque_num == cheque_number and 
                    excel_client.lower() == client_name.lower() and 
                    abs(excel_amount_float - cheque_amount) < 0.01)  # Small tolerance for float comparison
            
        except Exception as e:
            self.logger.error(f"Error checking exact match: {str(e)}")
            return False

    def _handle_duplicates_and_create_or_update(self, cheque, duplicates):
        """
        Handle case where duplicates exist - update the best match and remove others
        """
        try:
            if not duplicates:
                return self._create_new_cheque(cheque)
            
            # Find the best match (exact match preferred, otherwise first one)
            best_match = None
            for duplicate in duplicates:
                if duplicate['is_exact_match']:
                    best_match = duplicate
                    break
            
            if not best_match:
                best_match = duplicates[0]  # Use first one if no exact match
            
            # Update the best match
            success = self._update_excel_entry(cheque, best_match)
            
            if success:
                # Create mapping for the updated entry
                mapping = ChequeExcelMapping(
                    cheque_id=cheque.id,
                    excel_file_path=str(best_match['file_path']),
                    sheet_name=best_match['sheet_name'],
                    row_number=best_match['row_number']
                )
                db.session.add(mapping)
                
                # Remove all other duplicates
                other_duplicates = [d for d in duplicates if d != best_match]
                if other_duplicates:
                    self._remove_duplicate_entries(other_duplicates)
                
                db.session.commit()
                self.logger.info(f"Updated best match and removed {len(other_duplicates)} duplicates for cheque {cheque.id}")
            
            return success
            
        except Exception as e:
            self.logger.error(f"Error handling duplicates: {str(e)}")
            db.session.rollback()
            return False

    def _remove_duplicate_entries(self, duplicates, keep_mapping=None):
        """
        Remove duplicate entries from Excel files
        
        Args:
            duplicates: List of duplicate entries to remove
            keep_mapping: Optional mapping to keep (won't be removed)
        """
        try:
            # Group duplicates by file for efficient processing
            files_to_process = {}
            
            for duplicate in duplicates:
                # Skip if this is the mapping we want to keep
                if (keep_mapping and 
                    str(duplicate['file_path']) == keep_mapping.excel_file_path and
                    duplicate['sheet_name'] == keep_mapping.sheet_name and
                    duplicate['row_number'] == keep_mapping.row_number):
                    continue
                
                file_path = duplicate['file_path']
                if file_path not in files_to_process:
                    files_to_process[file_path] = {}
                
                sheet_name = duplicate['sheet_name']
                if sheet_name not in files_to_process[file_path]:
                    files_to_process[file_path][sheet_name] = []
                
                files_to_process[file_path][sheet_name].append(duplicate['row_number'])
            
            # Process each file
            for file_path, sheets in files_to_process.items():
                try:
                    workbook = load_workbook(file_path)
                    
                    for sheet_name, row_numbers in sheets.items():
                        if sheet_name in workbook.sheetnames:
                            worksheet = workbook[sheet_name]
                            
                            # Sort row numbers in descending order to maintain row indices when deleting
                            for row_num in sorted(row_numbers, reverse=True):
                                try:
                                    worksheet.delete_rows(row_num)
                                    self.logger.info(f"Removed duplicate from {sheet_name} row {row_num}")
                                except Exception as e:
                                    self.logger.error(f"Error deleting row {row_num}: {str(e)}")
                    
                    workbook.save(file_path)
                    workbook.close()
                    
                except Exception as e:
                    self.logger.error(f"Error processing file {file_path}: {str(e)}")
            
            self.logger.info(f"Completed removal of {len(duplicates)} duplicate entries")
            
        except Exception as e:
            self.logger.error(f"Error removing duplicates: {str(e)}")

    def _create_new_cheque(self, cheque):
        """Create new cheque entry in Excel with duplicate check"""
        try:
            # FINAL duplicate check before creating
            existing_duplicates = self._find_all_cheque_duplicates_in_excel(cheque)
            if existing_duplicates:
                self.logger.warning(f"Found duplicates during creation for cheque {cheque.id}, handling them first")
                return self._handle_duplicates_and_create_or_update(cheque, existing_duplicates)
            
            # Determine target file and sheet
            year = cheque.due_date.year
            month = cheque.due_date.month
            month_name = self.month_names[month - 1]
            
            # Get or create Excel file
            filepath = self.excel_folder / f"cheques_{year}.xlsx"
            workbook, worksheet = self._ensure_workbook_and_sheet(filepath, month_name)
            
            # Find next available row
            next_row = worksheet.max_row + 1
            
            # Write cheque data
            cheque_data = self._prepare_cheque_data(cheque)
            for col_idx, value in enumerate(cheque_data, 1):
                worksheet.cell(row=next_row, column=col_idx, value=value)
            
            # Apply formatting
            self._apply_formatting(worksheet, next_row)
            
            # Save workbook
            workbook.save(filepath)
            workbook.close()
            
            # Create mapping
            mapping = ChequeExcelMapping(
                cheque_id=cheque.id,
                excel_file_path=str(filepath),
                sheet_name=month_name,
                row_number=next_row
            )
            db.session.add(mapping)
            db.session.commit()
            
            self.logger.info(f"Created new cheque {cheque.id} in Excel at row {next_row}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error creating new cheque: {str(e)}")
            db.session.rollback()
            return False

    def _update_existing_cheque(self, cheque, mapping):
        """Update existing cheque in its original location"""
        try:
            filepath = Path(mapping.excel_file_path)
            if not filepath.exists():
                self.logger.warning(f"Excel file not found: {filepath}")
                return self._create_new_cheque(cheque)
            
            # Load workbook and sheet
            workbook = load_workbook(filepath)
            if mapping.sheet_name not in workbook.sheetnames:
                self.logger.warning(f"Sheet {mapping.sheet_name} not found in {filepath}")
                workbook.close()
                return self._create_new_cheque(cheque)
            
            worksheet = workbook[mapping.sheet_name]
            
            # Update the specific row
            cheque_data = self._prepare_cheque_data(cheque)
            for col_idx, value in enumerate(cheque_data, 1):
                worksheet.cell(row=mapping.row_number, column=col_idx, value=value)
            
            # Apply formatting
            self._apply_formatting(worksheet, mapping.row_number)
            
            # Save workbook
            workbook.save(filepath)
            workbook.close()
            
            # Update mapping timestamp
            mapping.updated_at = datetime.utcnow()
            db.session.commit()
            
            self.logger.info(f"Updated cheque {cheque.id} in {filepath}, sheet {mapping.sheet_name}, row {mapping.row_number}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error updating existing cheque: {str(e)}")
            db.session.rollback()
            return False

    def _update_excel_entry(self, cheque, excel_location):
        """Update existing Excel entry at specified location"""
        try:
            filepath = excel_location['file_path']
            sheet_name = excel_location['sheet_name']
            row_number = excel_location['row_number']
            
            workbook = load_workbook(filepath)
            worksheet = workbook[sheet_name]
            
            # Update the specific row
            cheque_data = self._prepare_cheque_data(cheque)
            for col_idx, value in enumerate(cheque_data, 1):
                worksheet.cell(row=row_number, column=col_idx, value=value)
            
            # Apply formatting
            self._apply_formatting(worksheet, row_number)
            
            # Save workbook
            workbook.save(filepath)
            workbook.close()
            
            self.logger.info(f"Updated cheque {cheque.id} in Excel at row {row_number}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error updating Excel entry: {str(e)}")
            return False

    def _handle_cheque_deletion(self, cheque):
        """Delete cheque from Excel and clean up all duplicates"""
        try:
            # Find ALL instances of this cheque in Excel
            all_duplicates = self._find_all_cheque_duplicates_in_excel(cheque)
            
            if all_duplicates:
                # Remove all instances
                self._remove_duplicate_entries(all_duplicates)
                self.logger.info(f"Deleted {len(all_duplicates)} instances of cheque {cheque.id} from Excel")
            
            # Clean up database mapping
            existing_mapping = ChequeExcelMapping.query.filter_by(cheque_id=cheque.id).first()
            if existing_mapping:
                db.session.delete(existing_mapping)
                db.session.commit()
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error deleting cheque: {str(e)}")
            db.session.rollback()
            return False

    def _validate_mapping_location(self, mapping):
        """Validate that the mapping location still exists"""
        try:
            filepath = Path(mapping.excel_file_path)
            if not filepath.exists():
                return False
            
            workbook = load_workbook(filepath, read_only=True)
            is_valid = mapping.sheet_name in workbook.sheetnames
            workbook.close()
            
            return is_valid
            
        except Exception as e:
            self.logger.error(f"Error validating mapping: {str(e)}")
            return False

    def _cleanup_invalid_mapping(self, mapping):
        """Remove invalid mapping from database"""
        try:
            db.session.delete(mapping)
            db.session.commit()
            self.logger.info(f"Cleaned up invalid mapping for cheque {mapping.cheque_id}")
        except Exception as e:
            self.logger.error(f"Error cleaning up mapping: {str(e)}")
            db.session.rollback()

    def _ensure_workbook_and_sheet(self, filepath, sheet_name):
        """Ensure workbook and sheet exist"""
        if filepath.exists():
            workbook = load_workbook(filepath)
        else:
            workbook = Workbook()
            # Remove default sheet
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
        
        if sheet_name not in workbook.sheetnames:
            worksheet = workbook.create_sheet(title=sheet_name)
            # Add headers
            for col_idx, header in enumerate(self.headers, 1):
                worksheet.cell(row=1, column=col_idx, value=header)
            self._apply_header_formatting(worksheet)
        else:
            worksheet = workbook[sheet_name]
        
        return workbook, worksheet

    def _prepare_cheque_data(self, cheque):
        """Prepare cheque data for Excel"""
        return [
            cheque.issue_date.strftime('%d/%m/%Y') if cheque.issue_date else '',
            cheque.payment_type or 'CHQ',
            cheque.cheque_number or '',
            f"{cheque.branch.bank.name} - {cheque.branch.name}" if cheque.branch else '',
            f"{cheque.deposit_branch.bank.name} - {cheque.deposit_branch.name}" if cheque.deposit_branch else '',
            cheque.client.name if cheque.client else '',
            cheque.depositor.name if cheque.depositor else cheque.depositor_name or '',
            float(cheque.amount),
            cheque.currency or 'MAD',
            cheque.due_date.strftime('%d/%m/%Y') if cheque.due_date else '',
            cheque.created_date.strftime('%d/%m/%Y') if cheque.created_date else '',
            cheque.status or 'EN ATTENTE',
            cheque.invoice_number or '',
            cheque.invoice_date.strftime('%d/%m/%Y') if cheque.invoice_date else '',
            cheque.notes or ''
        ]

    def _apply_header_formatting(self, worksheet):
        """Apply formatting to header row"""
        from openpyxl.styles import Font, Alignment, PatternFill
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col in range(1, len(self.headers) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    def _apply_formatting(self, worksheet, row_number):
        """Apply formatting to data row"""
        from openpyxl.styles import Alignment
        
        # Center align certain columns (dates, amounts, status)
        center_cols = [1, 2, 8, 9, 10, 11, 12, 13]  # Date and status columns
        for col in center_cols:
            cell = worksheet.cell(row=row_number, column=col)
            cell.alignment = Alignment(horizontal="center")

    def remove_all_duplicates_comprehensive(self, year=None):
        """
        COMPREHENSIVE removal of all duplicate entries in Excel files
        """
        try:
            if year is None:
                year = datetime.now().year
            
            filepath = self.excel_folder / f"cheques_{year}.xlsx"
            if not filepath.exists():
                return {'status': 'no_file', 'message': f'No Excel file for year {year}'}
            
            results = {
                'duplicates_found': 0,
                'duplicates_removed': 0,
                'sheets_processed': 0,
                'errors': []
            }
            
            workbook = load_workbook(filepath)
            
            for sheet_name in workbook.sheetnames:
                try:
                    worksheet = workbook[sheet_name]
                    sheet_results = self._remove_duplicates_from_sheet_comprehensive(worksheet, sheet_name)
                    results['duplicates_found'] += sheet_results['found']
                    results['duplicates_removed'] += sheet_results['removed']
                    results['sheets_processed'] += 1
                    results['errors'].extend(sheet_results['errors'])
                except Exception as e:
                    results['errors'].append(f"Error processing sheet {sheet_name}: {str(e)}")
            
            if results['duplicates_removed'] > 0:
                workbook.save(filepath)
                self.logger.info(f"Removed {results['duplicates_removed']} duplicates from {filepath}")
            
            workbook.close()
            return results
            
        except Exception as e:
            self.logger.error(f"Error removing duplicates: {str(e)}")
            return {'status': 'error', 'message': str(e)}

    def _remove_duplicates_from_sheet_comprehensive(self, worksheet, sheet_name):
        """COMPREHENSIVE duplicate removal from a specific sheet"""
        results = {'found': 0, 'removed': 0, 'errors': []}
        
        try:
            seen_cheques = {}  # Use dict to track first occurrence
            rows_to_delete = []
            
            # Check from row 2 (skip headers)
            for row_num in range(2, worksheet.max_row + 1):
                try:
                    cheque_number = worksheet.cell(row=row_num, column=3).value  # N° doc
                    
                    if not cheque_number:
                        continue
                    
                    cheque_key = str(cheque_number).strip()
                    
                    if cheque_key in seen_cheques:
                        # This is a duplicate
                        results['found'] += 1
                        rows_to_delete.append(row_num)
                        self.logger.info(f"Found duplicate cheque {cheque_key} in {sheet_name} at row {row_num} (first seen at row {seen_cheques[cheque_key]})")
                    else:
                        # First occurrence - remember it
                        seen_cheques[cheque_key] = row_num
                        
                except Exception as e:
                    results['errors'].append(f"Error checking row {row_num} in {sheet_name}: {str(e)}")
            
            # Delete duplicate rows (in reverse order to maintain row numbers)
            for row_num in reversed(rows_to_delete):
                try:
                    worksheet.delete_rows(row_num)
                    results['removed'] += 1
                    self.logger.info(f"Removed duplicate row {row_num} from {sheet_name}")
                except Exception as e:
                    results['errors'].append(f"Error deleting row {row_num} in {sheet_name}: {str(e)}")
            
        except Exception as e:
            results['errors'].append(f"Error processing sheet {sheet_name}: {str(e)}")
        
        return results

    # Additional utility methods remain the same...
    def verify_integrity_comprehensive(self, year=None):
        """Comprehensive integrity check"""
        try:
            from models import Cheque
            
            if year is None:
                year = datetime.now().year
            
            # Count cheques in database for the year
            db_cheques = Cheque.query.filter(
                db.extract('year', Cheque.due_date) == year
            ).all()
            
            # Count mappings for the year
            mapping_count = ChequeExcelMapping.query.filter(
                ChequeExcelMapping.excel_file_path.like(f'%cheques_{year}.xlsx')
            ).count()
            
            # Count actual entries in Excel and check for duplicates
            filepath = self.excel_folder / f"cheques_{year}.xlsx"
            excel_count = 0
            excel_cheques = []
            duplicate_cheques = []
            
            if filepath.exists():
                workbook = load_workbook(filepath, read_only=True)
                seen_cheques = set()
                
                for sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                    for row_num in range(2, worksheet.max_row + 1):
                        cheque_num = worksheet.cell(row=row_num, column=3).value
                        if cheque_num:
                            cheque_key = str(cheque_num).strip()
                            excel_count += 1
                            
                            if cheque_key in seen_cheques:
                                duplicate_cheques.append({
                                    'cheque_number': cheque_key,
                                    'sheet': sheet_name,
                                    'row': row_num
                                })
                            else:
                                seen_cheques.add(cheque_key)
                                
                            excel_cheques.append({
                                'cheque_number': cheque_key,
                                'sheet': sheet_name,
                                'row': row_num
                            })
                workbook.close()
            
            return {
                'year': year,
                'database_count': len(db_cheques),
                'mapping_count': mapping_count,
                'excel_count': excel_count,
                'excel_unique_count': len(seen_cheques) if 'seen_cheques' in locals() else 0,
                'excel_duplicates_count': len(duplicate_cheques),
                'db_excel_diff': abs(len(db_cheques) - excel_count),
                'db_mapping_diff': abs(len(db_cheques) - mapping_count),
                'excel_mapping_diff': abs(excel_count - mapping_count),
                'has_duplicates': len(duplicate_cheques) > 0,
                'in_sync': len(db_cheques) == mapping_count == len(seen_cheques) if 'seen_cheques' in locals() else False,
                'duplicate_cheques': duplicate_cheques,
                'excel_cheques': excel_cheques[:10]  # Sample for debugging
            }
            
        except Exception as e:
            self.logger.error(f"Error verifying integrity: {str(e)}")
            return {'error': str(e)}

    def batch_sync_all_cheques(self, cheques_query=None):
        """Batch synchronize all cheques with duplicate prevention"""
        from models import Cheque
        
        if cheques_query is None:
            cheques = Cheque.query.all()
        else:
            cheques = cheques_query.all()
        
        results = {
            'total': len(cheques),
            'successful': 0,
            'failed': 0,
            'duplicates_removed': 0,
            'errors': []
        }
        
        for cheque in cheques:
            try:
                if self.sync_cheque(cheque, 'create'):
                    results['successful'] += 1
                else:
                    results['failed'] += 1
            except Exception as e:
                results['failed'] += 1
                results['errors'].append(f"Cheque {cheque.id}: {str(e)}")
        
        self.logger.info(f"Batch sync completed: {results['successful']}/{results['total']} successful")
        return results