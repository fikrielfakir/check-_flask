"""
Advanced Excel Workbook Manager for yearly organization with 12 monthly sheets.
Optimized for offline desktop use with comprehensive cheque management.
"""

import os
import glob
from datetime import datetime, date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

class ExcelYearlyManager:
    """Manages per-year Excel workbooks, each with 12 monthly sheets"""
    
    def __init__(self, cheques_dir):
        """
        Initialize the Excel manager
        
        Args:
            cheques_dir (str): Directory where Excel files will be stored
        """
        self.cheques_dir = cheques_dir
        os.makedirs(self.cheques_dir, exist_ok=True)
        
        # Define column headers for consistency
        self.headers = [
            'Date', 'Type', 'Numéro', 'Banque', 'Propriétaire', 
            'Déposant', 'Montant', 'devise','Date échéance', 'Statut', 'Notes'
        ]
        
        # French month names
        self.month_names = [
            'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
            'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre'
        ]
        
        logging.info(f"Excel manager initialized for directory: {self.cheques_dir}")
    
    def create_yearly_file(self, year):
        """
        Create a new yearly Excel file with 12 monthly sheets
        
        Args:
            year (int): Year for the workbook
            
        Returns:
            str: Path to the created file
        """
        filename = f"cheques_{year}.xlsx"
        filepath = os.path.join(self.cheques_dir, filename)
        
        # Create workbook and remove default sheet
        wb = Workbook()
        wb.remove(wb.active)
        
        # Create 12 monthly sheets
        for month_num, month_name in enumerate(self.month_names, 1):
            ws = wb.create_sheet(title=month_name)
            self._setup_sheet_headers(ws)
        
        wb.save(filepath)
        logging.info(f"Created yearly file: {filepath}")
        return filepath
    
    def _setup_sheet_headers(self, worksheet):
        """
        Set up styled headers for a worksheet
        
        Args:
            worksheet: openpyxl worksheet object
        """
        # Updated headers to match the new format
        self.headers = [
            "Date d'émission",
            "Type de Règlement", 
            "Numéro du chèque",
            "Banque/Agence",
            "Client",
            "Nom du déposant",
            "Montant",
            "Devise",
            "Date d'échéance",
            "Date de Création",
            "Statut",
            "N° Facture",
            "Date de facture",
            "Notes"
        ]
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add headers
        for col_num, header in enumerate(self.headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Set column widths - Updated for new columns
        column_widths = {
            'A': 12,  # Date d'émission
            'B': 12,  # Type de Règlement
            'C': 15,  # Numéro du chèque
            'D': 25,  # Banque/Agence
            'E': 25,  # Client
            'F': 20,  # Nom du déposant
            'G': 12,  # Montant
            'H': 8,   # Devise
            'I': 12,  # Date d'échéance
            'J': 12,  # Date de Création
            'K': 12,  # Statut
            'L': 15,  # N° Facture
            'M': 12,  # Date de facture
            'N': 30   # Notes
        }
        
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
        
        # Freeze first row
        worksheet.freeze_panes = "A2"
    
    def ensure_file_and_sheet_exist(self, year, month):
        """
        Ensure yearly file and monthly sheet exist
        
        Args:
            year (int): Year
            month (int): Month (1-12)
            
        Returns:
            tuple: (workbook, worksheet)
        """
        filename = f"cheques_{year}.xlsx"
        filepath = os.path.join(self.cheques_dir, filename)
        
        # Create file if it doesn't exist
        if not os.path.exists(filepath):
            self.create_yearly_file(year)
        
        # Load workbook
        wb = load_workbook(filepath)
        
        # Get month sheet name
        month_name = self.month_names[month - 1]
        
        # Create sheet if it doesn't exist
        if month_name not in wb.sheetnames:
            ws = wb.create_sheet(title=month_name)
            self._setup_sheet_headers(ws)
            wb.save(filepath)
        else:
            ws = wb[month_name]
        
        return wb, ws
    
    def add_or_update_cheque(self, cheque_data):
        """
        Add or update a cheque in the appropriate yearly file and monthly sheet
        
        Args:
            cheque_data (dict): Cheque information including:
                - echeance_date: due date (datetime.date or string)
                - numero: cheque number
                - banque: bank name
                - propriétaire: owner name
                - deposant: depositor name
                - montant: amount
                - type: CHQ or LCN
                - statut: status
                - notes: notes
                - date_emission: issue date
                - devise: currency
                - date_creation: creation date
                - numero_facture: invoice number
                - date_facture: invoice date
        """
        try:
            # Parse due date to determine year and month
            if isinstance(cheque_data.get('echeance_date'), str):
                due_date = datetime.strptime(cheque_data['echeance_date'], '%Y-%m-%d').date()
            else:
                due_date = cheque_data.get('echeance_date', date.today())
            
            year = due_date.year
            month = due_date.month
            
            # Ensure file and sheet exist
            wb, ws = self.ensure_file_and_sheet_exist(year, month)
            
            # Check for existing cheque
            existing_row = self._find_existing_cheque(ws, cheque_data.get('numero'), cheque_data.get('banque'))
            
            # Prepare row data - Updated to match new format
            row_data = [
                cheque_data.get('date_emission', ''),
                cheque_data.get('type', 'CHQ'),
                cheque_data.get('numero', ''),
                cheque_data.get('banque', ''),
                cheque_data.get('propriétaire', ''),
                cheque_data.get('deposant', ''),
                cheque_data.get('montant', 0),
                cheque_data.get('devise', 'MAD'),
                cheque_data.get('echeance_date', ''),
                cheque_data.get('date_creation', ''),
                cheque_data.get('statut', 'EN_ATTENTE'),
                cheque_data.get('numero_facture', ''),
                cheque_data.get('date_facture', ''),
                cheque_data.get('notes', '')
            ]
            
            if existing_row:
                # Update existing row
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=existing_row, column=col_num, value=value)
                logging.info(f"Updated cheque {cheque_data.get('numero')} in row {existing_row}")
            else:
                # Find next empty row
                next_row = ws.max_row + 1
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=next_row, column=col_num, value=value)
                logging.info(f"Added new cheque {cheque_data.get('numero')} in row {next_row}")
            
            # Apply formatting
            self._format_worksheet(ws)
            
            # Save workbook
            filename = f"cheques_{year}.xlsx"
            filepath = os.path.join(self.cheques_dir, filename)
            wb.save(filepath)
            
            return True
            
        except Exception as e:
            logging.error(f"Error adding/updating cheque: {str(e)}")
            return False
    
    def _find_existing_cheque(self, worksheet, numero, banque):
        """
        Find existing cheque by number and bank
        
        Args:
            worksheet: openpyxl worksheet
            numero (str): cheque number
            banque (str): bank name
            
        Returns:
            int or None: Row number if found, None otherwise
        """
        for row in range(2, worksheet.max_row + 1):
            row_numero = worksheet.cell(row=row, column=3).value  # Column C (Numéro)
            row_banque = worksheet.cell(row=row, column=4).value  # Column D (Banque)
            
            if str(row_numero) == str(numero) and str(row_banque) == str(banque):
                return row
        
        return None
    
    def _format_worksheet(self, worksheet):
        """
        Apply formatting to worksheet (alternating row colors, borders)
        
        Args:
            worksheet: openpyxl worksheet
        """
        # Light blue fill for alternating rows
        light_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
        
        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply formatting to data rows
        for row in range(2, worksheet.max_row + 1):
            # Alternate row coloring
            if row % 2 == 0:
                for col in range(1, len(self.headers) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.fill = light_fill
            
            # Apply borders to all cells
            for col in range(1, len(self.headers) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = thin_border
                
                # Format amount column (G) as currency
                if col == 7:  # Montant column (G)
                    cell.number_format = '#,##0.00'
                
                # Center align date and status columns
                if col in [1, 2, 8, 9, 10, 11, 12, 13]:  # Date and status columns
                    cell.alignment = Alignment(horizontal="center")
    
    def remove_cheque(self, numero, banque, year=None):
        """
        Remove a cheque from the Excel files
        
        Args:
            numero (str): cheque number
            banque (str): bank name
            year (int, optional): specific year to search
            
        Returns:
            bool: True if removed, False if not found
        """
        if year:
            years = [year]
        else:
            # Search all years
            years = self.get_available_years()
        
        for year in years:
            filename = f"cheques_{year}.xlsx"
            filepath = os.path.join(self.cheques_dir, filename)
            
            if not os.path.exists(filepath):
                continue
            
            wb = load_workbook(filepath)
            
            for month_name in self.month_names:
                if month_name not in wb.sheetnames:
                    continue
                
                ws = wb[month_name]
                existing_row = self._find_existing_cheque(ws, numero, banque)
                
                if existing_row:
                    ws.delete_rows(existing_row)
                    wb.save(filepath)
                    logging.info(f"Removed cheque {numero} from {year}/{month_name}")
                    return True
        
        return False
    
    def get_file_info(self, year):
        """
        Get information about a yearly file
        
        Args:
            year (int): Year
            
        Returns:
            dict: File information including existence, path, sheets, counts, size, etc.
        """
        filename = f"cheques_{year}.xlsx"
        filepath = os.path.join(self.cheques_dir, filename)
        
        info = {
            'year': year,
            'filename': filename,
            'filepath': filepath,
            'exists': os.path.exists(filepath),
            'sheets': [],
            'total_cheques': 0,
            'file_size': 0,
            'modified_date': None
        }
        
        if info['exists']:
            try:
                # Get file stats
                stat = os.stat(filepath)
                info['file_size'] = stat.st_size
                info['modified_date'] = datetime.fromtimestamp(stat.st_mtime)
                
                # Load workbook to get sheet info
                wb = load_workbook(filepath, read_only=True)
                
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    cheque_count = max(0, ws.max_row - 1)  # Subtract header row
                    
                    info['sheets'].append({
                        'name': sheet_name,
                        'cheque_count': cheque_count
                    })
                    info['total_cheques'] += cheque_count
                
                wb.close()
                
            except Exception as e:
                logging.error(f"Error reading file info for {filepath}: {str(e)}")
        
        return info
    
    def list_all_files(self):
        """
        List all yearly Excel files in the directory
        
        Returns:
            list: List of file information dictionaries sorted by year
        """
        files = []
        pattern = os.path.join(self.cheques_dir, "cheques_*.xlsx")
        
        for filepath in glob.glob(pattern):
            filename = os.path.basename(filepath)
            try:
                year_str = filename.replace('cheques_', '').replace('.xlsx', '')
                year = int(year_str)
                files.append(self.get_file_info(year))
            except ValueError:
                logging.warning(f"Invalid year format in filename: {filename}")
        
        return sorted(files, key=lambda x: x['year'], reverse=True)
    
    def get_available_years(self):
        """
        Get list of available years
        
        Returns:
            list: Sorted list of years (integers)
        """
        files = self.list_all_files()
        return [file_info['year'] for file_info in files]
    
    def export_year_summary(self, year):
        """
        Export a yearly summary with statistics for each month
        
        Args:
            year (int): Year to export
            
        Returns:
            str: Path to the summary file
        """
        try:
            filename = f"cheques_{year}.xlsx"
            filepath = os.path.join(self.cheques_dir, filename)
            
            if not os.path.exists(filepath):
                raise FileNotFoundError(f"No file found for year {year}")
            
            # Create summary workbook
            summary_filename = f"resume_{year}.xlsx"
            summary_filepath = os.path.join(self.cheques_dir, summary_filename)
            
            summary_wb = Workbook()
            summary_ws = summary_wb.active
            summary_ws.title = f"Résumé {year}"
            
            # Headers for summary
            summary_headers = [
                'Mois', 'Nombre de chèques', 'Montant total', 'CHQ', 'LCN',
                'Encaissés', 'Impayés', 'En attente', 'Rejetés', 'Déposés', 'Annulés'
            ]
            
            # Set up headers
            for col_num, header in enumerate(summary_headers, 1):
                cell = summary_ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Load main workbook
            wb = load_workbook(filepath, read_only=True)
            
            total_row = ['TOTAL', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
            
            # Process each month
            for month_num, month_name in enumerate(self.month_names, 1):
                if month_name not in wb.sheetnames:
                    continue
                
                ws = wb[month_name]
                
                # Initialize counters
                month_stats = {
                    'count': 0,
                    'total_amount': 0,
                    'chq': 0,
                    'lcn': 0,
                    'encaisse': 0,
                    'impaye': 0,
                    'en_attente': 0,
                    'rejete': 0,
                    'depose': 0,
                    'annule': 0
                }
                
                # Process each row (skip header)
                for row in range(2, ws.max_row + 1):
                    if not ws.cell(row=row, column=1).value:  # Skip empty rows
                        continue
                    
                    month_stats['count'] += 1
                    
                    # Type (Column B)
                    cheque_type = str(ws.cell(row=row, column=2).value or '').upper()
                    if cheque_type == 'CHQ':
                        month_stats['chq'] += 1
                    elif cheque_type == 'LCN':
                        month_stats['lcn'] += 1
                    
                    # Amount (Column G)
                    amount = ws.cell(row=row, column=7).value or 0
                    try:
                        month_stats['total_amount'] += float(amount)
                    except (ValueError, TypeError):
                        pass
                    
                    # Status (Column I)
                    status = str(ws.cell(row=row, column=9).value or '').lower()
                    if 'encaisse' in status:
                        month_stats['encaisse'] += 1
                    elif 'impaye' in status:
                        month_stats['impaye'] += 1
                    elif 'attente' in status:
                        month_stats['en_attente'] += 1
                    elif 'rejete' in status:
                        month_stats['rejete'] += 1
                    elif 'depose' in status:
                        month_stats['depose'] += 1
                    elif 'annule' in status:
                        month_stats['annule'] += 1
                
                # Add month row to summary
                month_row = [
                    month_name,
                    month_stats['count'],
                    month_stats['total_amount'],
                    month_stats['chq'],
                    month_stats['lcn'],
                    month_stats['encaisse'],
                    month_stats['impaye'],
                    month_stats['en_attente'],
                    month_stats['rejete'],
                    month_stats['depose'],
                    month_stats['annule']
                ]
                
                for col_num, value in enumerate(month_row, 1):
                    summary_ws.cell(row=month_num + 1, column=col_num, value=value)
                
                # Update totals
                for i in range(1, len(total_row)):
                    if isinstance(month_row[i], (int, float)):
                        total_row[i] += month_row[i]
            
            # Add total row
            total_row_num = len(self.month_names) + 2
            for col_num, value in enumerate(total_row, 1):
                cell = summary_ws.cell(row=total_row_num, column=col_num, value=value)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Auto-size columns
            for column in summary_ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                summary_ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.close()
            summary_wb.save(summary_filepath)
            
            logging.info(f"Exported year summary: {summary_filepath}")
            return summary_filepath
            
        except Exception as e:
            logging.error(f"Error exporting year summary: {str(e)}")
            raise