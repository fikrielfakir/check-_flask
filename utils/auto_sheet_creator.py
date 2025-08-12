"""
Auto Sheet Creator - Optimizes Excel sheet creation
Automatically creates all 12 monthly sheets when adding a cheque to ensure complete yearly structure
"""

import os
import logging
from datetime import datetime
from openpyxl import Workbook, load_workbook
from pathlib import Path


class AutoSheetCreator:
    """
    Automatically ensures all monthly sheets exist when a cheque is added.
    This optimization prevents the need to create sheets one-by-one as cheques are added.
    """
    
    def __init__(self, excel_folder=None):
        self.excel_folder = Path(excel_folder) if excel_folder else Path("data/excel")
        self.excel_folder.mkdir(parents=True, exist_ok=True)
        
        # French month names for sheet naming
        self.month_names = [
            'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
            'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre'
        ]
        
        # Excel column headers for cheques
        self.headers = [
            'Date Émission', 'Type Paiement', 'N° chq', 'Banque - Agence',
            'Bq dép. - Agce', 'Client', 'Déposant', 'Mont', 'Devise',
            'Date d\'Échéance', 'Date Création', 'Statut', 'Numéro Facture',
            'Date Facture', 'Observations'
        ]

    def ensure_all_monthly_sheets(self, year):
        """
        Ensures all 12 monthly sheets exist for the given year.
        This is the main optimization - creates complete yearly structure at once.
        
        Args:
            year (int): The year for which to ensure all monthly sheets exist
            
        Returns:
            str: Path to the yearly Excel file
        """
        try:
            filename = f"cheques_{year}.xlsx"
            filepath = self.excel_folder / filename
            
            # Check if file exists
            if filepath.exists():
                # File exists, check if all monthly sheets are present
                workbook = load_workbook(str(filepath))
                existing_sheets = set(workbook.sheetnames)
                missing_sheets = set(self.month_names) - existing_sheets
                
                if missing_sheets:
                    # Create missing monthly sheets
                    for month_name in missing_sheets:
                        self._create_monthly_sheet(workbook, month_name)
                    
                    workbook.save(str(filepath))
                    logging.info(f"Added {len(missing_sheets)} missing monthly sheets to {filename}")
                else:
                    logging.debug(f"All monthly sheets already exist in {filename}")
                
                workbook.close()
            else:
                # Create new yearly file with all 12 monthly sheets
                workbook = Workbook()
                workbook.remove(workbook.active)  # Remove default sheet
                
                # Create all 12 monthly sheets
                for month_name in self.month_names:
                    self._create_monthly_sheet(workbook, month_name)
                
                workbook.save(str(filepath))
                workbook.close()
                logging.info(f"Created new yearly file with all 12 monthly sheets: {filename}")
            
            return str(filepath)
            
        except Exception as e:
            logging.error(f"Error ensuring monthly sheets for year {year}: {str(e)}")
            return None

    def _create_monthly_sheet(self, workbook, month_name):
        """
        Creates a single monthly sheet with proper headers and formatting.
        
        Args:
            workbook: OpenPyXL workbook object
            month_name (str): Name of the month for the sheet
        """
        try:
            worksheet = workbook.create_sheet(title=month_name)
            
            # Add headers
            for col_idx, header in enumerate(self.headers, 1):
                cell = worksheet.cell(row=1, column=col_idx, value=header)
                # Apply header formatting
                cell.font = cell.font.copy(bold=True)
                # Set column width based on content
                if col_idx <= len(self._get_column_widths()):
                    worksheet.column_dimensions[cell.column_letter].width = self._get_column_widths()[col_idx - 1]
                
        except Exception as e:
            logging.error(f"Error creating monthly sheet {month_name}: {str(e)}")

    def _get_column_widths(self):
        """Returns optimal column widths for cheque data"""
        return [12, 12, 15, 25, 25, 20, 20, 12, 8, 12, 12, 15, 15, 12, 30]

    def optimize_cheque_addition(self, cheque):
        """
        Main optimization function called when adding a cheque.
        Ensures all monthly sheets exist for the cheque's year.
        
        Args:
            cheque: Cheque object with due_date attribute
            
        Returns:
            tuple: (success: bool, filepath: str)
        """
        if not cheque or not hasattr(cheque, 'due_date') or not cheque.due_date:
            logging.error("Invalid cheque object or missing due_date")
            return False, None
            
        year = cheque.due_date.year
        
        try:
            filepath = self.ensure_all_monthly_sheets(year)
            
            if filepath:
                logging.info(f"Sheet optimization completed for cheque in year {year}")
                return True, filepath
            else:
                logging.error(f"Failed to optimize sheets for year {year}")
                return False, None
                
        except Exception as e:
            logging.error(f"Error in cheque addition optimization: {str(e)}")
            return False, None

    def batch_optimize_years(self, start_year, end_year=None):
        """
        Batch optimization for multiple years.
        Useful for initial setup or data migration.
        
        Args:
            start_year (int): Starting year
            end_year (int, optional): Ending year. Defaults to current year.
            
        Returns:
            dict: Summary of optimization results
        """
        if end_year is None:
            end_year = datetime.now().year
            
        results = {
            'success': 0,
            'failed': 0,
            'years_processed': [],
            'errors': []
        }
        
        for year in range(start_year, end_year + 1):
            try:
                filepath = self.ensure_all_monthly_sheets(year)
                if filepath:
                    results['success'] += 1
                    results['years_processed'].append(year)
                else:
                    results['failed'] += 1
                    results['errors'].append(f"Failed to process year {year}")
                    
            except Exception as e:
                results['failed'] += 1
                results['errors'].append(f"Error processing year {year}: {str(e)}")
        
        logging.info(f"Batch optimization completed: {results['success']} successful, {results['failed']} failed")
        return results

    def get_yearly_file_status(self, year):
        """
        Get the status of a yearly file and its monthly sheets.
        
        Args:
            year (int): Year to check
            
        Returns:
            dict: Status information
        """
        filename = f"cheques_{year}.xlsx"
        filepath = self.excel_folder / filename
        
        status = {
            'year': year,
            'file_exists': filepath.exists(),
            'filepath': str(filepath) if filepath.exists() else None,
            'monthly_sheets': [],
            'missing_sheets': [],
            'total_sheets': 0
        }
        
        if filepath.exists():
            try:
                workbook = load_workbook(str(filepath))
                existing_sheets = workbook.sheetnames
                status['monthly_sheets'] = existing_sheets
                status['total_sheets'] = len(existing_sheets)
                status['missing_sheets'] = [month for month in self.month_names if month not in existing_sheets]
                workbook.close()
                
            except Exception as e:
                status['error'] = str(e)
                logging.error(f"Error checking yearly file status for {year}: {str(e)}")
        
        return status


# Convenience function for easy integration
def auto_create_monthly_sheets(cheque, excel_folder=None):
    """
    Convenience function to automatically create all monthly sheets for a cheque's year.
    
    Args:
        cheque: Cheque object
        excel_folder (str, optional): Path to Excel folder
        
    Returns:
        tuple: (success: bool, filepath: str)
    """
    creator = AutoSheetCreator(excel_folder)
    return creator.optimize_cheque_addition(cheque)


# Integration hook for existing cheque creation workflow
def hook_into_cheque_creation(cheque_creation_function):
    """
    Decorator to automatically optimize sheet creation when adding cheques.
    
    Usage:
        @hook_into_cheque_creation
        def your_cheque_creation_function(cheque_data):
            # Your existing cheque creation logic
            return created_cheque
    """
    def wrapper(*args, **kwargs):
        # Call original function
        result = cheque_creation_function(*args, **kwargs)
        
        # If cheque was created successfully, optimize sheets
        if result and hasattr(result, 'due_date'):
            try:
                success, filepath = auto_create_monthly_sheets(result)
                if success:
                    logging.info(f"Auto-created monthly sheets for cheque {getattr(result, 'cheque_number', 'unknown')}")
                else:
                    logging.warning(f"Failed to auto-create monthly sheets for cheque {getattr(result, 'cheque_number', 'unknown')}")
            except Exception as e:
                logging.error(f"Error in sheet creation hook: {str(e)}")
        
        return result
    
    return wrapper