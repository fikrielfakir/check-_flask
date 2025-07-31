"""
Excel Manager routes for advanced yearly workbook management.
Integrates with the ExcelYearlyManager for comprehensive file operations.
"""

from flask import Blueprint, render_template, request, jsonify, flash, redirect, url_for, current_app, send_file
from flask_login import login_required, current_user
from datetime import datetime, date
import os
import logging
from utils.excel_yearly_manager import ExcelYearlyManager
from utils.database_manager import DatabaseManager
from models import Cheque, Branch, Client

excel_manager_bp = Blueprint('excel_manager', __name__)

@excel_manager_bp.route('/excel')
@login_required
def excel_dashboard():
    """Excel management dashboard"""
    try:
        excel_dir = current_app.config['EXCEL_FOLDER']
        excel_manager = ExcelYearlyManager(excel_dir)
        
        # Get all yearly files
        files_info = excel_manager.list_all_files()
        
        # Get current year file info
        current_year = datetime.now().year
        current_year_info = excel_manager.get_file_info(current_year)
        
        return render_template('excel/dashboard.html', 
                             files_info=files_info,
                             current_year_info=current_year_info,
                             current_year=current_year)
    
    except Exception as e:
        logging.error(f"Error in excel dashboard: {str(e)}")
        flash(f'Erreur lors du chargement du tableau de bord Excel: {str(e)}', 'error')
        return redirect(url_for('dashboard.index'))

@excel_manager_bp.route('/excel/create_year', methods=['POST'])
@login_required
def create_year_file():
    """Create a new yearly Excel file"""
    try:
        year = int(request.form.get('year', datetime.now().year))
        
        excel_dir = current_app.config['EXCEL_FOLDER']
        excel_manager = ExcelYearlyManager(excel_dir)
        
        # Check if file already exists
        file_info = excel_manager.get_file_info(year)
        if file_info['exists']:
            flash(f'Le fichier pour l\'année {year} existe déjà.', 'warning')
        else:
            # Create new yearly file
            file_path = excel_manager.create_yearly_file(year)
            flash(f'Fichier Excel créé avec succès pour l\'année {year}.', 'success')
            logging.info(f"Created yearly Excel file: {file_path}")
        
        return redirect(url_for('excel_manager.excel_dashboard'))
    
    except Exception as e:
        logging.error(f"Error creating yearly file: {str(e)}")
        flash(f'Erreur lors de la création du fichier: {str(e)}', 'error')
        return redirect(url_for('excel_manager.excel_dashboard'))

@excel_manager_bp.route('/excel/sync_database')
@login_required
def sync_database_to_excel():
    """Synchronize database cheques to Excel files"""
    try:
        excel_dir = current_app.config['EXCEL_FOLDER']
        from utils.cheque_excel_sync import ChequeExcelSync
        
        sync_manager = ChequeExcelSync(excel_dir)
        
        # Get all cheques from database
        cheques = Cheque.query.all()
        
        # Perform bulk synchronization
        sync_results = sync_manager.bulk_sync_all_cheques()
        
        if sync_results['failed_syncs'] == 0:
            flash(f'Synchronisation réussie: {sync_results["successful_syncs"]} chèques synchronisés.', 'success')
        else:
            flash(f'Synchronisation partielle: {sync_results["successful_syncs"]} réussies, {sync_results["failed_syncs"]} erreurs.', 'warning')
        
        return redirect(url_for('excel_manager.excel_dashboard'))
    
    except Exception as e:
        logging.error(f"Error syncing database to Excel: {str(e)}")
        flash(f'Erreur lors de la synchronisation: {str(e)}', 'error')
        return redirect(url_for('excel_manager.excel_dashboard'))

@excel_manager_bp.route('/excel/export_year_summary/<int:year>')
@login_required
def export_year_summary(year):
    """Export yearly summary report"""
    try:
        excel_dir = current_app.config['EXCEL_FOLDER']
        excel_manager = ExcelYearlyManager(excel_dir)
        
        # Generate summary
        summary_path = excel_manager.export_year_summary(year)
        
        # Send file for download
        return send_file(summary_path, 
                        as_attachment=True,
                        download_name=f'resume_{year}.xlsx',
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    except Exception as e:
        logging.error(f"Error exporting year summary: {str(e)}")
        flash(f'Erreur lors de l\'export du résumé: {str(e)}', 'error')
        return redirect(url_for('excel_manager.excel_dashboard'))

@excel_manager_bp.route('/excel/file_details/<int:year>')
@login_required
def file_details(year):
    """Get detailed information about a yearly file"""
    try:
        excel_dir = current_app.config['EXCEL_FOLDER']
        excel_manager = ExcelYearlyManager(excel_dir)
        
        file_info = excel_manager.get_file_info(year)
        
        return jsonify({
            'success': True,
            'data': file_info
        })
    
    except Exception as e:
        logging.error(f"Error getting file details: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        })

@excel_manager_bp.route('/excel/download/<int:year>')
@login_required
def download_yearly_file(year):
    """Download yearly Excel file"""
    try:
        excel_dir = current_app.config['EXCEL_FOLDER']
        filename = f'cheques_{year}.xlsx'
        file_path = os.path.join(excel_dir, filename)
        
        if not os.path.exists(file_path):
            flash(f'Le fichier pour l\'année {year} n\'existe pas.', 'error')
            return redirect(url_for('excel_manager.excel_dashboard'))
        
        return send_file(file_path, 
                        as_attachment=True,
                        download_name=filename,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    except Exception as e:
        logging.error(f"Error downloading file: {str(e)}")
        flash(f'Erreur lors du téléchargement: {str(e)}', 'error')
        return redirect(url_for('excel_manager.excel_dashboard'))

@excel_manager_bp.route('/excel/statistics')
@login_required
def excel_statistics():
    """Get Excel files statistics"""
    try:
        excel_dir = current_app.config['EXCEL_FOLDER']
        excel_manager = ExcelYearlyManager(excel_dir)
        
        # Get all files info
        files_info = excel_manager.list_all_files()
        
        # Calculate statistics
        stats = {
            'total_files': len(files_info),
            'total_cheques': sum(file_info['total_cheques'] for file_info in files_info),
            'years_covered': [file_info['year'] for file_info in files_info],
            'total_size': sum(file_info['file_size'] for file_info in files_info),
            'files_by_year': {file_info['year']: file_info for file_info in files_info}
        }
        
        return jsonify({
            'success': True,
            'statistics': stats
        })
    
    except Exception as e:
        logging.error(f"Error getting Excel statistics: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        })

@excel_manager_bp.route('/excel/add_cheque', methods=['POST'])
@login_required
def add_cheque_to_excel():
    """Add a single cheque directly to Excel"""
    try:
        excel_dir = current_app.config['EXCEL_FOLDER']
        excel_manager = ExcelYearlyManager(excel_dir)
        
        # Get form data
        cheque_data = {
            'date_emission': request.form.get('date_emission'),
            'type': request.form.get('type', 'CHQ'),
            'numero': request.form.get('numero'),
            'banque': request.form.get('banque'),
            'proprietaire': request.form.get('proprietaire'),
            'deposant': request.form.get('deposant'),
            'montant': float(request.form.get('montant', 0)),
            'echeance_date': datetime.strptime(request.form.get('echeance_date'), '%Y-%m-%d').date(),
            'statut': request.form.get('statut', 'EN_ATTENTE'),
            'notes': request.form.get('notes', '')
        }
        
        # Add to Excel
        if excel_manager.add_or_update_cheque(cheque_data):
            flash('Chèque ajouté avec succès dans Excel.', 'success')
        else:
            flash('Erreur lors de l\'ajout du chèque dans Excel.', 'error')
        
        return redirect(url_for('excel_manager.excel_dashboard'))
    
    except Exception as e:
        logging.error(f"Error adding cheque to Excel: {str(e)}")
        flash(f'Erreur lors de l\'ajout: {str(e)}', 'error')
        return redirect(url_for('excel_manager.excel_dashboard'))

@excel_manager_bp.route('/excel/remove_cheque', methods=['POST'])
@login_required
def remove_cheque_from_excel():
    """Remove a cheque from Excel files"""
    try:
        excel_dir = current_app.config['EXCEL_FOLDER']
        excel_manager = ExcelYearlyManager(excel_dir)
        
        numero = request.form.get('numero')
        banque = request.form.get('banque')
        year = request.form.get('year')
        
        if year:
            year = int(year)
        else:
            year = None
        
        # Remove from Excel
        if excel_manager.remove_cheque(numero, banque, year):
            flash('Chèque supprimé avec succès d\'Excel.', 'success')
        else:
            flash('Chèque non trouvé dans les fichiers Excel.', 'warning')
        
        return redirect(url_for('excel_manager.excel_dashboard'))
    
    except Exception as e:
        logging.error(f"Error removing cheque from Excel: {str(e)}")
        flash(f'Erreur lors de la suppression: {str(e)}', 'error')
        return redirect(url_for('excel_manager.excel_dashboard'))

@excel_manager_bp.route('/excel/backup_files')
@login_required
def backup_excel_files():
    """Create backup of all Excel files"""
    try:
        excel_dir = current_app.config['EXCEL_FOLDER']
        backup_dir = os.path.join(current_app.config['EXPORTS_FOLDER'], 'backups')
        os.makedirs(backup_dir, exist_ok=True)
        
        import shutil
        import zipfile
        from datetime import datetime
        
        # Create backup filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_filename = f'excel_backup_{timestamp}.zip'
        backup_path = os.path.join(backup_dir, backup_filename)
        
        # Create zip file with all Excel files
        with zipfile.ZipFile(backup_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(excel_dir):
                for file in files:
                    if file.endswith('.xlsx'):
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, excel_dir)
                        zipf.write(file_path, arcname)
        
        flash(f'Sauvegarde créée avec succès: {backup_filename}', 'success')
        
        return send_file(backup_path, 
                        as_attachment=True,
                        download_name=backup_filename,
                        mimetype='application/zip')
    
    except Exception as e:
        logging.error(f"Error creating backup: {str(e)}")
        flash(f'Erreur lors de la création de la sauvegarde: {str(e)}', 'error')
        return redirect(url_for('excel_manager.excel_dashboard'))

@excel_manager_bp.route('/excel/cleanup')
@login_required
def cleanup_excel_files():
    """Clean up empty or corrupted Excel files"""
    try:
        excel_dir = current_app.config['EXCEL_FOLDER']
        excel_manager = ExcelYearlyManager(excel_dir)
        
        # Get all files
        files_info = excel_manager.list_all_files()
        
        cleaned_count = 0
        for file_info in files_info:
            if file_info['exists'] and file_info['total_cheques'] == 0:
                try:
                    # Check if it's truly empty (only headers)
                    from openpyxl import load_workbook
                    wb = load_workbook(file_info['filepath'], read_only=True)
                    
                    is_empty = True
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        if ws.max_row > 1:  # More than just header
                            is_empty = False
                            break
                    
                    wb.close()
                    
                    if is_empty:
                        # Remove empty file
                        os.remove(file_info['filepath'])
                        cleaned_count += 1
                        logging.info(f"Removed empty Excel file: {file_info['filepath']}")
                
                except Exception as e:
                    logging.warning(f"Error checking file {file_info['filepath']}: {str(e)}")
        
        if cleaned_count > 0:
            flash(f'Nettoyage terminé: {cleaned_count} fichiers vides supprimés.', 'success')
        else:
            flash('Aucun fichier à nettoyer trouvé.', 'info')
        
        return redirect(url_for('excel_manager.excel_dashboard'))
    
    except Exception as e:
        logging.error(f"Error during cleanup: {str(e)}")
        flash(f'Erreur lors du nettoyage: {str(e)}', 'error')
        return redirect(url_for('excel_manager.excel_dashboard'))

@excel_manager_bp.route('/excel/open_folder')
@login_required
def open_excel_folder():
    """Open Excel folder in file explorer (Windows)"""
    try:
        excel_dir = current_app.config['EXCEL_FOLDER']
        
        # For Windows
        import subprocess
        import platform
        
        if platform.system() == 'Windows':
            subprocess.run(['explorer', excel_dir], check=True)
            flash('Dossier Excel ouvert dans l\'explorateur.', 'success')
        else:
            flash('Fonctionnalité disponible uniquement sous Windows.', 'warning')
        
        return redirect(url_for('excel_manager.excel_dashboard'))
    
    except Exception as e:
        logging.error(f"Error opening folder: {str(e)}")
        flash(f'Erreur lors de l\'ouverture du dossier: {str(e)}', 'error')
        return redirect(url_for('excel_manager.excel_dashboard'))